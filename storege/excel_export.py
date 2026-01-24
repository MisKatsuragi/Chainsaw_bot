import sys
import os
from openpyxl import Workbook
from .data_manager import DataManager, dm
from .databases.items_db import ItemsDatabase, Item

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def ExcelExport(dm: DataManager, path: str = "database_export.xlsx"):  # ✅ Принимает dm
    """Экспорт всех данных в Excel"""
    wb = Workbook()
    
    # Рынок
    ws_market = wb.active
    ws_market.title = "Рынок"
    ws_market.append(["ID", "Название", "Цена", "Описание"])
    for item in dm.market_items:
        ws_market.append([item.index, item.name, item.cost, item.description])
    
    # Пользователи
    ws_users = wb.create_sheet("Пользователи")
    ws_users.append(["ID", "Монеты", "Должность", "Получено", "Потрачено"])
    for user_id, user_data in dm.users.items():
        stats = user_data.stats
        ws_users.append([
            user_id, user_data.coins, stats.position,
            stats.total_received, stats.total_spent
        ])
    
    # Инвентарь
    ws_inventory = wb.create_sheet("Инвентарь")
    ws_inventory.append(["UserID", "Слот", "ItemID", "Название", "Цена"])
    for user_id, user_data in dm.users.items():
        for slot, item in user_data.items.items():
            ws_inventory.append([user_id, slot, item.index, item.name, item.cost])
    
    wb.save(path)
    return f"✅ Экспорт завершён: {path}"